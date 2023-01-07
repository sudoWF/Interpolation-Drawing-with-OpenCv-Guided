import sys
from openpyxl import Workbook
from openpyxl import load_workbook

# --------------------从txt中提取数据--------------------
# --------------------转换格式到txt中--------------------
result = []
with open('points.txt', 'r') as f:
    for line in f:
        result.append(list(line.strip('\n').split(',')))
# print(result)
# print(result[0] == ['{'])

# -------------------把数据放到excel里-------------------
workbook = Workbook()
final_sheet = workbook.create_sheet('final',0)
temp_sheet = workbook.create_sheet('temp',1)
# 每个workbook创建后，默认会存在一个worksheet，对默认的worksheet进行重命名
# worksheet.title = "Sheet1"
for row in result:
    temp_sheet.append(row)  # 把每一行append到worksheet中
workbook.save(filename="points.xlsx")  # 不能忘

print("Txt to Excel has Finished")

# -----------------在新sheet里将各轮廓分开-----------------
# workbook1 = load_workbook(filename="points.xlsx")
# sheet = workbook["Sheet1"]
# data = sheet.cell(row=1, column=1)
# print("单元格的值:", data.value, "\n单元格的行数:", data.row, "\n单元格的列数:", data.column, "\n单元格的坐标:", data.coordinate)
# data = temp_sheet.cell(row=1, column=1)
# print(data.value)
# print(data.value=='{')
# print(temp_sheet.max_row)

# final_sheet.cell(row=2, column=2).value = data.value

# i = 0
# print([i,data.value])

n=0
i=1
while i < temp_sheet.max_row+1 :
    data1 = temp_sheet.cell(row=i, column=1)
    data2 = temp_sheet.cell(row=i, column=2)
    if data1.value == '{':
        i += 1
        # pass
    elif data1.value == '}':
        i += 1
        n+=1
    else:
        final_sheet.append([n, data1.value, data2.value])
        # print([n, data1.value, data2.value])
        i += 1

print('Distinguishing the data has completed')

# --------------------转换格式到txt中--------------------
f=open('final.txt','w')

i = 1
while i < final_sheet.max_row+1 :
    num = final_sheet.cell(row=i, column=1)
    x_point = final_sheet.cell(row=i, column=2)
    y_point = final_sheet.cell(row=i, column=3)
    result = '{' + str(num.value) + ',' + str(x_point.value) + ',' + str(y_point.value) + '},\n'
    f.write(result)
    i += 1

f.close()

print('Result txt has completed')



workbook.save(filename="points.xlsx")  # 不能忘