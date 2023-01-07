import cv2
import numpy as np

import sys
import re

import sys
from openpyxl import Workbook
from openpyxl import load_workbook


# --------------------提取图片中的轮廓--------------------
image = cv2.imread("000.png")
gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)  # 转为灰度图
blurred = cv2.GaussianBlur(gray, (11, 11), 0)
edge = cv2.Canny(blurred, 30, 150)  # 用Canny算子提取边缘
#cv2.imwrite("edge.jpg", edge)
cv2.imshow("image_1", edge)
cv2.waitKey(0)

contour = image.copy()
(cnts, _) = cv2.findContours(edge, cv2.RETR_LIST, cv2.CHAIN_APPROX_SIMPLE)  # 轮廓检测
cv2.drawContours(contour, cnts, -1, (0, 255, 0), 2)  # 绘制轮廓
# cv2.imwrite("./results/contour.jpg", contour)
cv2.imshow("image_2", contour)
cv2.waitKey(0)

np.savetxt("points.txt",cnts,fmt='%s',delimiter=',')
print('0-图像识别已完成...')



# ------------------去除轮廓数据中的空行-------------------
# 删除空行并生成新文件
with open('points.txt', 'r+') as fr, open('temp.txt', 'w+') as fd:
    for text in fr.readlines():
        if text.split():
            fd.write(text)
fr.close()
fd.close()
print('1-已去除空行...')



# ------------------轮廓数据去除多重括号-------------------
f1 = open('temp.txt','r',encoding='UTF-8')
# f1 = open('points.txt', 'r+')
f2 = open('points.txt', 'w+')
# print(f1.read())
# before = f1.read()
# middle = before.replace('[[[','{\n[').replace(']]]',']\n}').replace(' [[','[').replace(']]',']').replace(' ',',')
# after = middle.replace('[,','[').replace(',,',',')
before = f1.read()
middle = before.replace('[[[ ','{\n').replace('[[[','{\n').replace(']]]','\n}').replace(' [[ ','').replace(' [[','').replace(']]','').replace(' ',',')
after = middle.replace(',,',',')
# final = after.replace('\n',',')

# print(after)

# a = after.split("\n")
# print(a)
f2.write(after)
f1.close()
f2.close()

print('2-已合并括号...')



# --------------------从txt中提取数据--------------------
# -------------------在Excel中规范数据-------------------
# 从txt中提取数据
result = []
with open('points.txt', 'r') as f:
    for line in f:
        result.append(list(line.strip('\n').split(',')))
# print(result)
# print(result[0] == ['{'])

# 把数据放到excel里
workbook = Workbook()
final_sheet = workbook.create_sheet('final',0)
temp_sheet = workbook.create_sheet('temp',1)
# 每个workbook创建后，默认会存在一个worksheet，对默认的worksheet进行重命名
# worksheet.title = "Sheet1"
for row in result:
    temp_sheet.append(row)  # 把每一行append到worksheet中
workbook.save(filename="points.xlsx")  # 不能忘

# print("Txt to Excel has Finished")

# 在新sheet里将各轮廓分开
n=0
i=1
while i < temp_sheet.max_row+1 :
    data1 = temp_sheet.cell(row=i, column=1)
    data2 = temp_sheet.cell(row=i, column=2)
    if data1.value == '{':
        i += 1
        # pass
    elif data1.value == '}':
        i += 1
        n+=1
    else:
        final_sheet.append([n, data1.value, data2.value])
        # print([n, data1.value, data2.value])
        i += 1

print(n)
print('3-已规范数据...')

# --------------------转换格式到txt中--------------------
f=open('final.txt','w')

i = 1
while i < final_sheet.max_row+1 :
    num = final_sheet.cell(row=i, column=1)
    x_point = final_sheet.cell(row=i, column=2)
    y_point = final_sheet.cell(row=i, column=3)
    result = '{' + str(num.value) + ',' + str(x_point.value) + ',' + str(y_point.value) + '},\n'
    f.write(result)
    i += 1

f.close()

print('4-已输出数据...')


workbook.save(filename="points.xlsx")  # 不能忘



# ---------------------合并生成C代码---------------------
# 坐标点总数
count = len(open("final.txt",encoding="utf-8").readlines())
# print(count)
# print(type(count))

contours = n #总程序里的轮廓数

c_file = open('cHello.txt','w')

for line in open('num_1.txt', encoding="utf-8"):
    c_file.writelines(line)
# c_file.write('\n')

c_file.writelines(str(count))

c_file.writelines('][3]={\n')

for line in open('final.txt', encoding="utf-8"):
    c_file.writelines(line)
c_file.write('\n')

for line in open('num_2.txt', encoding="utf-8"):
    c_file.writelines(line)
# c_file.write('\n')

c_file.writelines(str(contours))

for line in open('num_3.txt', encoding="utf-8"):
    c_file.writelines(line)
c_file.write('\n')

c_file.close()

print('5-已输出C程序为cHello.txt...')
print('Finished')
