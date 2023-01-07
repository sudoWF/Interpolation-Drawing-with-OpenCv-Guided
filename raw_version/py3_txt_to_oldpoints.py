import sys
from openpyxl import Workbook
from openpyxl import load_workbook

# --------------------从txt中提取数据--------------------
# --------------------转换格式到txt中--------------------
result = []
with open('points.txt', 'r') as f:
    for line in f:
        result.append(list(line.strip('\n').split(',')))
# print(result)
# print(result[0] == ['{'])

# -------------------把数据放到excel里-------------------
workbook = Workbook()
final_sheet = workbook.create_sheet('final',0)
temp_sheet = workbook.create_sheet('temp',1)
# 每个workbook创建后，默认会存在一个worksheet，对默认的worksheet进行重命名
# worksheet.title = "Sheet1"
for row in result:
    temp_sheet.append(row)  # 把每一行append到worksheet中
workbook.save(filename="points.xlsx")  # 不能忘

print("Txt to Excel has Finished")

# -----------------在新sheet里将各轮廓分开-----------------
# workbook1 = load_workbook(filename="points.xlsx")
# sheet = workbook["Sheet1"]
# data = sheet.cell(row=1, column=1)
# print("单元格的值:", data.value, "\n单元格的行数:", data.row, "\n单元格的列数:", data.column, "\n单元格的坐标:", data.coordinate)
# data = temp_sheet.cell(row=1, column=1)
# print(data.value)
# print(data.value=='{')
# print(temp_sheet.max_row)

# final_sheet.cell(row=2, column=2).value = data.value

# i = 0
# print([i,data.value])

n=0
i=1
while i < temp_sheet.max_row+1 :
    data1 = temp_sheet.cell(row=i, column=1)
    data2 = temp_sheet.cell(row=i, column=2)
    if data1.value == '{':
        i += 1
        # pass
    elif data1.value == '}':
        i += 1
        n+=1
    else:
        final_sheet.append([n, data1.value, data2.value])
        # print([n, data1.value, data2.value])
        i += 1

print('Distinguishing the data has completed')

workbook.save(filename="points.xlsx")  # 不能忘

# --------------------转换格式到txt中--------------------
# openpyxl里sheet.max_row可以获取最大行，但是这个变量有个问题，就是对那些原先有数据，后来又删除的表，容易出现获取最大行不是所需要值的情况，这是因为openpyxl把一些格式改变过的单元格也算成有效行了。
# 所以写了一个函数来获取真正有效的最大数据行.

# import openpyxl
# filename = 'test_max_row.xlsx'
# wb=openpyxl.load_workbook(filename)
# ws = wb.active
# i= ws.max_row
# print("max_row获得的最大行是：",i)
# print(type(ws.cell(1,2).value))

def get_max_row(sheet):
    i=sheet.max_row
    real_max_row = 0
    while i > 0:
        row_dict = {i.value for i in sheet[i]}
        if row_dict == {None}:
            i = i-1
        else:
            real_max_row = i
            break

    return real_max_row

# j = get_max_row(ws)
# print("通过自定义函数获取到的最大行是：", j)

# --------------------转换格式到txt中--------------------
f=open('final.txt','w')

maxrow = get_max_row(final_sheet)
i = 1
t = 1
# r = 0
# print(final_sheet.max_row)
f.write('u16 point[4][6]={\n')
while i < maxrow+1 :
    num1 = final_sheet.cell(row=i, column=1)
    num2 = final_sheet.cell(row=i+1, column=1)
    x_point1 = final_sheet.cell(row=i, column=2)
    x_point2 = final_sheet.cell(row=i+1, column=2)
    y_point1 = final_sheet.cell(row=i, column=3)
    y_point2 = final_sheet.cell(row=i+1, column=3)
    if num1.value == num2.value:
        result = '{1,' + str(x_point1.value) + ',' + str(y_point1.value) + ',' + str(x_point2.value) + ',' + str(y_point2.value) + ',0},\n'
        f.write(result)
    else:
        # result = '};\nu16 point' + str(t) + '[4][6]={\n'
        # t += 1
        pass
#     print(num1.value,num2.value,'\n')    
    i += 1

f.write('};')
f.close()

print('Result txt has completed')



workbook.save(filename="points.xlsx")  # 不能忘
